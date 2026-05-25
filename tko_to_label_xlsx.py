"""
TKOファイルからラベル.xlsxを生成するスクリプト

ｼｽﾃﾑ管理シートが複数（ｼｽﾃﾑ管理、ｼｽﾃﾑ管理2、ｼｽﾃﾑ管理3…）あっても自動で処理します。
処理前にカテゴリ数とカテゴリ名の個数（空文字除外）の整合性を検証し、
不一致があれば警告を表示しつつ処理を続行し、エラーレポートをテキストファイルに出力します。

使い方:
    python tko_to_label_xlsx.py <入力TKOファイル> <出力xlsxファイル>

例:
    python tko_to_label_xlsx.py ラベル.TKO ラベル.xlsx

必要ライブラリ:
    pip install xlrd openpyxl
"""

import sys
import os
from datetime import datetime
from functools import lru_cache
import xlrd
import openpyxl
from copy import copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Color, Side
from openpyxl.utils import get_column_letter


TYPE_MAP   = {'ＳＡ': 'SA', 'ＭＡ': 'MA', '数量': 'NU', '文字': 'TX'}
COLUMNS    = ['列№', 'ラベル', '質問タイトル', 'タイプ', 'カテゴリ数', 'カテゴリ№', 'カテゴリ']
INT_COLS   = {'列№', 'カテゴリ数', 'カテゴリ№'}

# 出力シートの列幅
COL_WIDTHS = {
    'A': 5.7109375,
    'B': 15.7109375,
    'C': 35.7109375,
    'D': 5.42578125,
    'E': 4.28515625,
    'F': None,        # 未指定（デフォルト）
    'G': 70.7109375,
}

# ヘッダー書式
HEADER_FONT      = Font(name='ＭＳ ゴシック', size=9)
HEADER_FILL      = PatternFill(fill_type='solid', fgColor=Color(theme=9, tint=0.7999511703848384))
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT        = Font(name='ＭＳ ゴシック', size=9)
GRID_SIDE        = Side(style='hair', color=Color(indexed=64))
NO_SIDE          = Side()
HEADER_ROW_HEIGHT = 33.75


def validate_categories(wb: xlrd.Book, output_path: str) -> bool:
    """
    全ｼｽﾃﾑ管理シートのカテゴリ数と実際のカテゴリ名個数を検証する。

    検証内容：
      - カテゴリ数（行10）と実際のカテゴリ名の個数（空文字除外）が一致しているか
      - カテゴリ数を超えてカテゴリ名が存在していないか

    不一致があれば警告を表示し、エラーレポートを出力ファイルと同じフォルダに出力する。
    処理は続行する。

    Returns:
        True: すべて一致 / False: 不一致あり（警告・レポート出力済み）
    """
    warnings = []

    for sheet_name in wb.sheet_names():
        if not sheet_name.startswith('ｼｽﾃﾑ管理'):
            continue
        ws = wb.sheet_by_name(sheet_name)

        for c in range(1, ws.ncols):
            item_no = ws.cell_value(5, c)
            if item_no == '':
                break
            label = str(ws.cell_value(6, c)).strip()
            if not label:
                continue
            n_cats_val = ws.cell_value(10, c)
            if n_cats_val == '' or int(n_cats_val) == 0:
                continue

            n_cats = int(n_cats_val)

            actual_count = sum(
                1 for i in range(n_cats)
                if str(ws.cell_value(15 + i, c)).strip() != ''
            )

            extra_count = 0
            for r in range(15 + n_cats, min(15 + n_cats + 10, ws.nrows)):
                if str(ws.cell_value(r, c)).strip() != '':
                    extra_count += 1
                else:
                    break

            if actual_count != n_cats:
                warnings.append(
                    f"  [{sheet_name}] 項目No.{int(item_no)} ({label}): "
                    f"カテゴリ数={n_cats}, 実際のカテゴリ名={actual_count}個"
                )

            if extra_count > 0:
                warnings.append(
                    f"  [{sheet_name}] 項目No.{int(item_no)} ({label}): "
                    f"カテゴリ数={n_cats} だがさらに {extra_count} 個余分なカテゴリ名あり"
                )

    if warnings:
        print(
            f"\n【警告】カテゴリ数とカテゴリ名の個数が一致しない項目が {len(warnings)} 件あります。"
            "\n処理は続行しますが、該当項目を確認してください。\n"
            + "\n".join(warnings) + "\n"
        )

        timestamp   = datetime.now().strftime("%m%d%H%M")
        out_dir     = os.path.dirname(os.path.abspath(output_path))
        out_base    = os.path.splitext(os.path.basename(output_path))[0]
        report_path = os.path.join(out_dir, f"error_report_{out_base}_{timestamp}.txt")

        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("エラーレポート\n")
            f.write(f"生成日時    : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"出力ファイル: {os.path.abspath(output_path)}\n")
            f.write(f"{'='*60}\n\n")
            f.write(f"【警告】カテゴリ数とカテゴリ名の個数が一致しない項目が {len(warnings)} 件あります。\n\n")
            f.write("\n".join(warnings) + "\n")

        print(f"エラーレポートを出力しました: {report_path}")
        return False

    print("検証OK: カテゴリ数とカテゴリ名の個数はすべて一致しています。")
    return True


def process_sheet(ws: xlrd.sheet.Sheet, a_col_counter: int) -> tuple:
    """
    ｼｽﾃﾑ管理シート1枚分を処理する。

    Args:
        ws:             xlrd のシートオブジェクト
        a_col_counter:  列№ の開始値（複数シートをまたぐ場合は引き継ぐ）

    Returns:
        (rows, a_col_counter):
            rows           — 変換済み行データのリスト
            a_col_counter  — 次のシートに渡すカウンター値
    """
    labels     = ws.row_values(6)
    types_raw  = ws.row_values(8)
    cat_counts = ws.row_values(10)
    questions  = ws.row_values(12)

    rows = []
    for c in range(1, ws.ncols):
        item_no = ws.cell_value(5, c)
        if item_no == '':
            break
        item_no = int(item_no)

        label = str(labels[c]).strip()
        if not label:
            continue

        tipo     = TYPE_MAP.get(types_raw[c], types_raw[c])
        n_cats   = int(cat_counts[c]) if cat_counts[c] != '' else None
        question = str(questions[c]).strip()
        a_col_val = a_col_counter

        if n_cats and n_cats > 0:
            if tipo == 'MA':
                for cat_idx in range(1, n_cats + 1):
                    cat_val = str(ws.cell_value(14 + cat_idx, c)).strip()
                    rows.append({
                        '列№'       : a_col_val + cat_idx - 1,
                        'ラベル'    : f"{label}_{cat_idx}",
                        '質問タイトル': question if cat_idx == 1 else None,
                        'タイプ'    : tipo      if cat_idx == 1 else None,
                        'カテゴリ数' : n_cats    if cat_idx == 1 else None,
                        'カテゴリ№' : cat_idx,
                        'カテゴリ'  : cat_val,
                    })
                a_col_counter += n_cats
            else:
                for cat_idx in range(1, n_cats + 1):
                    cat_val = str(ws.cell_value(14 + cat_idx, c)).strip()
                    rows.append({
                        '列№'       : a_col_val if cat_idx == 1 else None,
                        'ラベル'    : label     if cat_idx == 1 else None,
                        '質問タイトル': question if cat_idx == 1 else None,
                        'タイプ'    : tipo      if cat_idx == 1 else None,
                        'カテゴリ数' : n_cats    if cat_idx == 1 else None,
                        'カテゴリ№' : cat_idx,
                        'カテゴリ'  : cat_val,
                    })
                a_col_counter += 1
        else:
            rows.append({
                '列№'       : a_col_val,
                'ラベル'    : label,
                '質問タイトル': question,
                'タイプ'    : tipo,
                'カテゴリ数' : None,
                'カテゴリ№' : None,
                'カテゴリ'  : None,
            })
            a_col_counter += 1

    return rows, a_col_counter


@lru_cache(maxsize=None)
def make_grid_border(has_left: bool, has_right: bool, has_top: bool, has_bottom: bool) -> Border:
    """見本と同じく、表の内側だけに細い罫線を引く。"""
    return Border(
        left=GRID_SIDE if has_left else NO_SIDE,
        right=GRID_SIDE if has_right else NO_SIDE,
        top=GRID_SIDE if has_top else NO_SIDE,
        bottom=GRID_SIDE if has_bottom else NO_SIDE,
    )


def tko_to_xlsx(input_path: str, output_path: str) -> None:
    """
    TKO ファイルを読み込み、ラベル.xlsx 形式で出力する。
    ｼｽﾃﾑ管理シートが複数ある場合は自動的にすべて処理し、列№を通しで採番する。
    処理前にカテゴリ数の整合性を検証し、不一致があれば警告とエラーレポートを出力して続行する。

    Args:
        input_path:  入力 TKO ファイルのパス
        output_path: 出力 xlsx ファイルのパス
    """
    wb = xlrd.open_workbook(input_path, encoding_override='cp932')

    # ── 1. カテゴリ数の整合性検証 ────────────────────────────────────
    print("カテゴリ数を検証中...")
    validate_categories(wb, output_path)

    # ── 2. ｼｽﾃﾑ管理系シートを順番に処理 ────────────────────────────
    all_rows = []
    a_col_counter = 1

    target_sheets = [
        name for name in wb.sheet_names()
        if name.startswith('ｼｽﾃﾑ管理')
    ]

    for sheet_name in target_sheets:
        ws = wb.sheet_by_name(sheet_name)

        has_items = any(
            str(ws.cell_value(6, c)).strip()
            for c in range(1, ws.ncols)
            if ws.cell_value(5, c) != ''
        )
        if not has_items:
            print(f"  {sheet_name}: 項目なし（スキップ）")
            continue

        rows, a_col_counter = process_sheet(ws, a_col_counter)
        all_rows.extend(rows)
        print(f"  {sheet_name}: {len(rows)} 行処理済み")

    # ── 3. xlsx 書き出し ─────────────────────────────────────────────
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'ラベル'
    max_row = len(all_rows) + 1
    max_col = len(COLUMNS)

    # ヘッダー行
    out_ws.row_dimensions[1].height = HEADER_ROW_HEIGHT
    for i, h in enumerate(COLUMNS, 1):
        cell = out_ws.cell(row=1, column=i, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = copy(HEADER_FILL)
        cell.alignment = HEADER_ALIGNMENT
        cell.border    = make_grid_border(i > 1, i < max_col, False, max_row > 1)

    # 列幅
    for col_letter, width in COL_WIDTHS.items():
        if width is not None:
            out_ws.column_dimensions[col_letter].width = width

    # データ行
    for row_idx, row in enumerate(all_rows, 2):
        out_row = []
        for col_name in COLUMNS:
            val = row[col_name]
            if col_name in INT_COLS and val is not None:
                val = int(val)
            out_row.append(val)
        out_ws.append(out_row)

        has_bottom = row_idx < max_row
        for col_idx in range(1, max_col + 1):
            cell = out_ws.cell(row_idx, col_idx)
            cell.font = DATA_FONT
            cell.border = make_grid_border(
                col_idx > 1,
                col_idx < max_col,
                True,
                has_bottom,
            )

    out_wb.save(output_path)
    print(f"出力完了: {output_path}  (合計 {len(all_rows)} 行, 最終列№={a_col_counter - 1})")


def resolve_paths(args: list[str]) -> tuple[str, str]:
    script_dir = os.path.dirname(os.path.abspath(__file__))

    if len(args) == 0:
        tko_files = [
            os.path.join(script_dir, name)
            for name in os.listdir(script_dir)
            if name.lower().endswith('.tko')
        ]
        if len(tko_files) == 1:
            input_path = tko_files[0]
            output_path = os.path.splitext(input_path)[0] + '.xlsx'
            return input_path, output_path

        if len(tko_files) == 0:
            raise ValueError("同じフォルダに .TKO ファイルが見つかりません。")

        names = "\n".join(f"  - {os.path.basename(path)}" for path in tko_files)
        raise ValueError(
            "同じフォルダに .TKO ファイルが複数あります。入力ファイルを指定してください。\n"
            + names
        )

    if len(args) == 1:
        input_path = args[0]
        output_path = os.path.splitext(input_path)[0] + '.xlsx'
        return input_path, output_path

    if len(args) == 2:
        return args[0], args[1]

    raise ValueError("使い方: python tko_to_label_xlsx.py [入力TKOファイル] [出力xlsxファイル]")


def run_gui() -> int:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    script_dir = os.path.dirname(os.path.abspath(__file__))
    root = tk.Tk()
    root.withdraw()

    input_path = filedialog.askopenfilename(
        title='TKOファイルを選択してください',
        initialdir=script_dir,
        filetypes=[('TKO files', '*.TKO'), ('All files', '*.*')],
    )
    if not input_path:
        return 0

    default_output = os.path.splitext(input_path)[0] + '.xlsx'
    output_path = filedialog.asksaveasfilename(
        title='出力xlsxファイルを指定してください',
        initialdir=os.path.dirname(default_output),
        initialfile=os.path.basename(default_output),
        defaultextension='.xlsx',
        filetypes=[('Excel workbook', '*.xlsx'), ('All files', '*.*')],
    )
    if not output_path:
        return 0

    try:
        print(f"入力: {input_path}")
        print(f"出力: {output_path}")
        tko_to_xlsx(input_path, output_path)
    except Exception as exc:
        messagebox.showerror('TKO to Label', f"変換に失敗しました。\n{exc}")
        return 1

    messagebox.showinfo('TKO to Label', f"変換が完了しました。\n{output_path}")
    return 0


def main(args: list[str]) -> int:
    if len(args) == 1 and args[0].lower() in {'--gui', '/gui'}:
        return run_gui()

    try:
        input_file, output_file = resolve_paths(args)
        print(f"入力: {input_file}")
        print(f"出力: {output_file}")
        tko_to_xlsx(input_file, output_file)
    except Exception as exc:
        print(f"エラー: {exc}")
        return 1

    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
